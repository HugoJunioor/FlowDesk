import csv, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

src = r"C:\Users\hugoc\Documents\op-es-suaves\redes_integrador.csv"
dst = r"C:\Users\hugoc\Documents\op-es-suaves\redes_integrador_x_relatorio.xlsx"

rows = []
with open(src, encoding="utf-8") as f:
    reader = csv.reader(f, delimiter="|")
    header = next(reader)
    for r in reader:
        if len(r) == 3:
            rows.append([r[0], r[1], int(r[2])])

wb = openpyxl.Workbook()

ws1 = wb.active
ws1.title = "Resumo por Relatorio"
ws1.append(["rel_nome", "qtd_redes"])
totals = {}
for integ, rel, q in rows:
    totals[rel] = totals.get(rel, 0) + q
for rel, q in sorted(totals.items(), key=lambda x: -x[1]):
    ws1.append([rel, q])
ws1.append(["TOTAL", sum(totals.values())])

ws2 = wb.create_sheet("Integrador x Relatorio")
ws2.append(["crd_red_nome_integrador", "rel_nome", "qtd_redes"])
for r in sorted(rows, key=lambda x: (-x[2], x[0])):
    ws2.append(r)

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="305496")
for ws in (ws1, ws2):
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

wb.save(dst)
print("OK ->", dst, "linhas detalhadas:", len(rows))
